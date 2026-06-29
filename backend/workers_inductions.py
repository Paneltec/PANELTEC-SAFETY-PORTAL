"""Phase 3.11 — Live Inductions Matrix.

Endpoints (all under `/api/workers/inductions`):
  POST /import-xlsx          — preview only (no writes); returns matched /
                                unmatched rows + per-cell confidence
  POST /import-xlsx/commit   — persists reviewed rows to
                                `worker_certifications` (with
                                `category` ∈ {site_induction, competency,
                                license} and `column_key`) and
                                `worker_access`
  GET  /matrix               — live matrix payload (rows = workers,
                                cols = distinct `(column_key, header,
                                category)`, cells = cert state + status)
  PUT  /cell                 — single-cell edit (set date, toggle not_held,
                                clear, attach existing doc_file_id)
  GET  /export.xlsx          — round-trip the live matrix back out as
                                an .xlsx (Inductions sheet + Access sheet)

Lenient date parser policy — confirmed by user (skip and flag, no silent
best-guesses):
  high          → ISO-like, full-year, unambiguous (e.g. 10/9/2020, 2024-06-15)
  medium        → AU DD/MM/YY with all parts in-range (06/24 → 1 Jun 2024)
  low           → ambiguous month/year, AND we still return date=None so the
                   wizard surfaces it for admin fix-up
  unparseable   → garbage / impossible / completely empty
The wizard groups low+unparseable under "Needs review"; commit writes only
high+medium. Low-confidence cells that survive end up as `invalid_date`
status in the matrix with a "Fix" affordance.
"""
from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime, timezone
from typing import Optional

import httpx
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel

from db import db
from auth import get_current_user

log = logging.getLogger("paneltec.workers.inductions")
router = APIRouter(prefix="/workers/inductions", tags=["workers-inductions"])


# ────────────────── lenient date parser ──────────────────

_NOT_HELD = {"no", "n", "nil", "none", "n/a", "na", "-", "--", "not held", "false"}
_HELD_OK  = {"yes", "y", "true", "ok", "current", "held"}

# Strip leading words like "Exp", "Expires", "Issued", "Renewed"
_LABEL = re.compile(r"^(?:exp(?:ires|iry)?|due|issued|renewed)\s*[:\-]?\s*", re.I)


def parse_messy_date(raw) -> dict:
    """Parse a free-text induction cell into {date, confidence, reason?,
    not_held?, raw}. Never raises — unparseable inputs return
    confidence='unparseable'."""
    if raw is None:
        return {"date": None, "confidence": "unparseable", "raw": "", "reason": "empty"}
    s_raw = str(raw).strip()
    if not s_raw:
        return {"date": None, "confidence": "unparseable", "raw": "", "reason": "empty"}
    s = s_raw

    # Explicit "not held" markers.
    if s.lower() in _NOT_HELD:
        return {"date": None, "confidence": "high", "raw": s_raw, "not_held": True}
    # Held without date — medium (we know they have it, just no expiry).
    if s.lower() in _HELD_OK:
        return {"date": None, "confidence": "medium", "raw": s_raw, "held": True,
                "reason": "no_expiry_date"}

    # Already a datetime/date object (openpyxl converts native Excel dates).
    if isinstance(raw, datetime):
        return {"date": raw.date().isoformat(), "confidence": "high", "raw": s_raw}
    if isinstance(raw, date):
        return {"date": raw.isoformat(), "confidence": "high", "raw": s_raw}

    s = _LABEL.sub("", s).strip()

    # ISO-like — high.
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            try:
                return {"date": date(y, mo, d).isoformat(), "confidence": "high", "raw": s_raw}
            except ValueError:
                return {"date": None, "confidence": "unparseable", "raw": s_raw, "reason": "impossible_value"}

    # DD/MM/YYYY or D/M/YYYY — high (AU convention is unambiguous with 4-digit year).
    m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            try:
                return {"date": date(y, mo, d).isoformat(), "confidence": "high", "raw": s_raw}
            except ValueError:
                return {"date": None, "confidence": "unparseable", "raw": s_raw, "reason": "impossible_value"}
        return {"date": None, "confidence": "unparseable", "raw": s_raw,
                "reason": "impossible_value"}

    # DD/MM/YY — medium (AU convention, 2-digit year resolved to 2000+yy or 1900+yy heuristic).
    m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            yyyy = 2000 + y if y <= 60 else 1900 + y
            try:
                return {"date": date(yyyy, mo, d).isoformat(), "confidence": "medium", "raw": s_raw}
            except ValueError:
                return {"date": None, "confidence": "unparseable", "raw": s_raw, "reason": "impossible_value"}
        # Out-of-range month/day (e.g. 23/22/24) — explicit reject, no silent flip.
        return {"date": None, "confidence": "unparseable", "raw": s_raw,
                "reason": "impossible_value"}

    # MM/YY or M/YY (e.g. "06/24", "5/24") — LOW confidence: no day → ambiguous
    # whether this means "expires end-of-month" or "issued on the 1st". User
    # decision: do not silently guess.
    m = re.match(r"^(\d{1,2})[/\-](\d{2})$", s)
    if m:
        mo, y = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            return {"date": None, "confidence": "low", "raw": s_raw,
                    "reason": "ambiguous_month_only",
                    "month_hint": mo, "year_hint": 2000 + y if y <= 60 else 1900 + y}
        return {"date": None, "confidence": "unparseable", "raw": s_raw,
                "reason": "impossible_value"}

    # Common typo: digit run with embedded slash like "10/1124" (= 10/11/24 likely).
    m = re.match(r"^(\d{1,2})[/\-](\d{2})(\d{2})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return {"date": None, "confidence": "low", "raw": s_raw,
                    "reason": "typo_concatenated_year",
                    "guess": date(2000 + y, mo, d).isoformat() if y <= 60 else date(1900 + y, mo, d).isoformat()}
        return {"date": None, "confidence": "unparseable", "raw": s_raw,
                "reason": "impossible_value"}

    return {"date": None, "confidence": "unparseable", "raw": s_raw, "reason": "no_pattern_match"}


# ────────────────── helpers ──────────────────

# Map column-header heuristics → category + canonical column_key.
_CATEGORY_HINTS = [
    ("site_induction", ["taswater", "cdo", "mv", "nmc", "tas rail", "tasrail", "wtc",
                         "ajr", "downer", "dcc", "airport", "fh"]),
    ("competency",     ["first aid", "cpr", "white card", "traffic control", "tma"]),
    ("license",        ["licen", "mr ", "hr "]),
]


def _category_for_header(h: str) -> str:
    hl = (h or "").lower()
    for cat, keys in _CATEGORY_HINTS:
        if any(k in hl for k in keys):
            return cat
    return "competency"  # heuristic default


def _column_key(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (h or "").lower()).strip("_") or "col"


def _norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").lower().strip())


# ────────────────── preview endpoint ──────────────────

class ImportIn(BaseModel):
    url: str


@router.post("/import-xlsx")
async def import_xlsx_preview(body: ImportIn, user: dict = Depends(get_current_user)):
    """Fetch + parse an inductions .xlsx. Returns matched/unmatched workers,
    per-cell confidence, and column → category mapping. **Does NOT save.**
    Caller must POST /commit with the reviewed payload."""
    if user.get("role") not in {"admin", "manager", "hseq_lead"}:
        raise HTTPException(403, "Only admin/manager/HSEQ can import inductions")

    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as c:
        r = await c.get(body.url)
        if r.status_code >= 400:
            raise HTTPException(400, f"Could not fetch xlsx ({r.status_code})")
        content = r.content

    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    # Load workers once and build a fuzzy-name index.
    workers_idx: dict[str, dict] = {}
    async for w in db.workers.find(
        {"org_id": user["org_id"], "deleted_at": None},
        {"_id": 0, "id": 1, "first_name": 1, "last_name": 1, "name": 1, "email": 1},
    ):
        full = w.get("name") or " ".join(filter(None, [w.get("first_name"), w.get("last_name")])).strip()
        if full:
            workers_idx[_norm_name(full)] = w
            # Also index "Last, First" and "First Last" rearrangements.
            last_first = f"{w.get('last_name','')} {w.get('first_name','')}".strip()
            workers_idx[_norm_name(last_first)] = w

    def match_worker(name_raw: str) -> Optional[dict]:
        k = _norm_name(name_raw)
        if not k: return None
        if k in workers_idx: return workers_idx[k]
        # Loose fuzzy: any worker whose full name shares ≥2 tokens with the input.
        tokens = set(k.split())
        if len(tokens) < 2: return None
        for ik, w in workers_idx.items():
            if len(set(ik.split()) & tokens) >= 2:
                return w
        return None

    out: dict = {
        "sheets_present": wb.sheetnames,
        "inductions": {"columns": [], "matched_rows": [], "unmatched_rows": [],
                        "low_confidence_cells": []},
        "accessibilities": {"matched_rows": [], "unmatched_rows": []},
    }

    # ── Sheet 1: Inductions ──
    sheet_name = next((s for s in wb.sheetnames if "induction" in s.lower()), wb.sheetnames[0])
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return out

    # Find header row: pick the first row containing >= 3 non-empty cells.
    header_i = next((i for i, r in enumerate(rows) if sum(1 for c in r if c) >= 3), 0)
    headers = [str(c).strip() if c is not None else "" for c in rows[header_i]]
    cols = []
    for i, h in enumerate(headers[1:], start=1):
        if not h:
            continue
        cols.append({
            "index": i, "header": h,
            "column_key": _column_key(h),
            "category": _category_for_header(h),
        })
    out["inductions"]["columns"] = cols

    for r_i, row in enumerate(rows[header_i + 1:], start=header_i + 2):
        worker_name = (row[0] or "")
        if not str(worker_name).strip():
            continue
        worker = match_worker(str(worker_name))
        cells = []
        for col in cols:
            val = row[col["index"]] if col["index"] < len(row) else None
            parsed = parse_messy_date(val)
            cells.append({
                "column_key": col["column_key"],
                "header": col["header"],
                "category": col["category"],
                **parsed,
            })
            if parsed["confidence"] in {"low", "unparseable"} and (val not in (None, "")):
                out["inductions"]["low_confidence_cells"].append({
                    "row": r_i, "worker_name": str(worker_name),
                    "matched_worker_id": worker.get("id") if worker else None,
                    "column": col["header"], "raw": parsed["raw"],
                    "reason": parsed.get("reason"),
                })
        bucket = "matched_rows" if worker else "unmatched_rows"
        out["inductions"][bucket].append({
            "row": r_i, "raw_name": str(worker_name),
            "worker_id": worker["id"] if worker else None,
            "worker_name": (
                worker.get("name") or
                " ".join(filter(None, [worker.get("first_name"), worker.get("last_name")])).strip()
            ) if worker else None,
            "cells": cells,
        })

    # ── Sheet 2: Accessibilities ──
    if any("access" in s.lower() for s in wb.sheetnames):
        s2 = wb[next(s for s in wb.sheetnames if "access" in s.lower())]
        rows = list(s2.iter_rows(values_only=True))
        if rows:
            header_i = next((i for i, r in enumerate(rows) if sum(1 for c in r if c) >= 2), 0)
            for row in rows[header_i + 1:]:
                if not row or not row[0]: continue
                name = str(row[0])
                worker = match_worker(name)
                # Naive column positions (vehicle, building_key, gate_key, extras).
                def _truthy(v):
                    s = str(v or "").strip().lower()
                    return s in {"y", "yes", "true", "1"}
                rec = {
                    "row": rows.index(row) + 1, "raw_name": name,
                    "worker_id": worker["id"] if worker else None,
                    "vehicle":      _truthy(row[1] if len(row) > 1 else None),
                    "building_key": _truthy(row[2] if len(row) > 2 else None),
                    "gate_key":     _truthy(row[3] if len(row) > 3 else None),
                    "extras":       str(row[4]) if len(row) > 4 and row[4] is not None else "",
                }
                bucket = "matched_rows" if worker else "unmatched_rows"
                out["accessibilities"][bucket].append(rec)

    out["summary"] = {
        "matched_workers": len(out["inductions"]["matched_rows"]),
        "unmatched_workers": len(out["inductions"]["unmatched_rows"]),
        "induction_columns": len(cols),
        "low_confidence_cells": len(out["inductions"]["low_confidence_cells"]),
        "access_matched": len(out["accessibilities"]["matched_rows"]),
        "access_unmatched": len(out["accessibilities"]["unmatched_rows"]),
    }
    return out


# ────────────────── commit endpoint ──────────────────

class CommitCell(BaseModel):
    column_key: str
    header: str
    category: str
    date: Optional[str] = None         # ISO YYYY-MM-DD or null
    not_held: Optional[bool] = False
    held: Optional[bool] = False
    confidence: Optional[str] = "high"


class CommitRow(BaseModel):
    worker_id: str
    cells: list[CommitCell]


class CommitAccessRow(BaseModel):
    worker_id: str
    vehicle: bool = False
    building_key: bool = False
    gate_key: bool = False
    extras: str = ""


class CommitIn(BaseModel):
    inductions: list[CommitRow] = []
    access: list[CommitAccessRow] = []


def _new_id() -> str:
    import uuid
    return uuid.uuid4().hex[:24]


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


@router.post("/import-xlsx/commit")
async def import_xlsx_commit(body: CommitIn, user: dict = Depends(get_current_user)):
    """Write the reviewed-and-confirmed rows. Only `high` / `medium`
    confidence cells with a date should be sent — wizard filters."""
    if user.get("role") not in {"admin", "manager", "hseq_lead"}:
        raise HTTPException(403, "Only admin/manager/HSEQ can commit inductions")

    inserted = 0; updated = 0; skipped = 0
    access_upserts = 0
    org = user["org_id"]

    for row in body.inductions:
        for cell in row.cells:
            # Skip cells with no information at all.
            if not (cell.date or cell.not_held or cell.held):
                skipped += 1
                continue
            existing = await db.worker_certifications.find_one(
                {"org_id": org, "worker_id": row.worker_id,
                 "column_key": cell.column_key, "deleted_at": None},
                {"_id": 0, "id": 1},
            )
            doc = {
                "org_id": org, "worker_id": row.worker_id,
                "name": cell.header,
                "category": cell.category,
                "column_key": cell.column_key,
                "expiry_date": cell.date,
                "not_held": bool(cell.not_held),
                "held_no_expiry": bool(cell.held) and not cell.date,
                "source": "induction_xlsx",
                "import_confidence": cell.confidence,
                "updated_at": _now_iso(),
                "updated_by": user["id"],
            }
            if existing:
                await db.worker_certifications.update_one(
                    {"id": existing["id"]}, {"$set": doc},
                )
                updated += 1
            else:
                doc.update({
                    "id": _new_id(), "issuer": "", "issue_date": None,
                    "doc_file_id": None, "doc_folder_id": None,
                    "notes": "", "created_by": user["id"],
                    "created_at": _now_iso(), "deleted_at": None,
                })
                await db.worker_certifications.insert_one(doc)
                inserted += 1

    for ar in body.access:
        await db.worker_access.update_one(
            {"org_id": org, "worker_id": ar.worker_id},
            {"$set": {
                "org_id": org, "worker_id": ar.worker_id,
                "vehicle": ar.vehicle, "building_key": ar.building_key,
                "gate_key": ar.gate_key, "extras": ar.extras,
                "updated_at": _now_iso(), "updated_by": user["id"],
            }, "$setOnInsert": {"id": _new_id(), "created_at": _now_iso()}},
            upsert=True,
        )
        access_upserts += 1

    return {
        "ok": True,
        "certifications": {"inserted": inserted, "updated": updated, "skipped": skipped},
        "access": {"upserted": access_upserts},
    }


# ────────────────── matrix GET ──────────────────

def _status_for(cert: dict) -> str:
    """Return one of: current | expiring | expired | not_held | held_no_expiry
    | invalid_date | unknown."""
    if cert.get("not_held"):
        return "not_held"
    exp = cert.get("expiry_date")
    if not exp:
        return "held_no_expiry" if cert.get("held_no_expiry") else "unknown"
    try:
        d = date.fromisoformat(exp)
    except Exception:
        return "invalid_date"
    today = date.today()
    if d < today:
        return "expired"
    delta_days = (d - today).days
    if delta_days <= 60:
        return "expiring"
    return "current"


@router.get("/matrix")
async def induction_matrix(user: dict = Depends(get_current_user)):
    """Live matrix payload. The frontend renders rows × columns with status
    chips and inline-edit affordances."""
    org = user["org_id"]
    # Workers list (active only).
    workers: list[dict] = []
    async for w in db.workers.find(
        {"org_id": org, "deleted_at": None},
        {"_id": 0, "id": 1, "first_name": 1, "last_name": 1, "name": 1,
         "email": 1, "company": 1, "status": 1},
    ).sort([("last_name", 1), ("first_name", 1), ("name", 1)]):
        workers.append({
            "id": w["id"],
            "name": w.get("name") or " ".join(filter(None,
                [w.get("first_name"), w.get("last_name")])).strip() or w.get("email") or "—",
            "email": w.get("email") or "",
            "company": w.get("company") or "",
            "status": w.get("status") or "active",
        })

    # Distinct induction columns from worker_certifications (source=induction_xlsx
    # OR rows that have a column_key set via a manual cell edit).
    pipeline = [
        {"$match": {"org_id": org, "deleted_at": None,
                    "column_key": {"$exists": True, "$ne": None}}},
        {"$group": {"_id": {"k": "$column_key", "h": "$name", "c": "$category"}}},
        {"$sort": {"_id.c": 1, "_id.h": 1}},
    ]
    cols = []
    async for row in db.worker_certifications.aggregate(pipeline):
        cols.append({
            "column_key": row["_id"]["k"],
            "header": row["_id"]["h"],
            "category": row["_id"]["c"] or "competency",
        })

    # Cells.
    cells_by_worker: dict[str, dict] = {}
    async for c in db.worker_certifications.find(
        {"org_id": org, "deleted_at": None,
         "column_key": {"$exists": True, "$ne": None}},
        {"_id": 0},
    ):
        cells_by_worker.setdefault(c["worker_id"], {})[c["column_key"]] = {
            "cert_id": c["id"],
            "expiry_date": c.get("expiry_date"),
            "not_held": bool(c.get("not_held")),
            "held_no_expiry": bool(c.get("held_no_expiry")),
            "doc_file_id": c.get("doc_file_id"),
            "status": _status_for(c),
            "import_confidence": c.get("import_confidence"),
        }

    # Access table.
    access_by_worker: dict[str, dict] = {}
    async for a in db.worker_access.find(
        {"org_id": org}, {"_id": 0},
    ):
        access_by_worker[a["worker_id"]] = {
            "vehicle": bool(a.get("vehicle")),
            "building_key": bool(a.get("building_key")),
            "gate_key": bool(a.get("gate_key")),
            "extras": a.get("extras") or "",
        }

    # Row-level summary chip.
    def _worker_chip(cells: dict) -> str:
        if not cells: return "unknown"
        statuses = {v["status"] for v in cells.values()}
        if "expired" in statuses: return "expired"
        if "invalid_date" in statuses: return "invalid_date"
        if "expiring" in statuses: return "expiring"
        if "unknown" in statuses and not (statuses & {"current", "held_no_expiry", "not_held"}):
            return "unknown"
        return "current"

    rows = []
    for w in workers:
        wc = cells_by_worker.get(w["id"], {})
        rows.append({
            **w,
            "cells": wc,
            "access": access_by_worker.get(w["id"], {"vehicle": False, "building_key": False, "gate_key": False, "extras": ""}),
            "chip": _worker_chip(wc),
        })

    return {"columns": cols, "rows": rows,
            "summary": {"workers": len(rows), "columns": len(cols)}}


# ────────────────── cell PUT ──────────────────

class CellPatch(BaseModel):
    worker_id: str
    column_key: str
    header: Optional[str] = None
    category: Optional[str] = None
    expiry_date: Optional[str] = None
    not_held: Optional[bool] = None
    held_no_expiry: Optional[bool] = None
    clear: Optional[bool] = False
    doc_file_id: Optional[str] = None


@router.put("/cell")
async def edit_cell(body: CellPatch, user: dict = Depends(get_current_user)):
    """Inline edit of a single matrix cell. Upserts the underlying
    `worker_certifications` document keyed by (worker_id, column_key)."""
    if user.get("role") not in {"admin", "manager", "hseq_lead"}:
        raise HTTPException(403, "Only admin/manager/HSEQ can edit inductions")

    org = user["org_id"]
    existing = await db.worker_certifications.find_one(
        {"org_id": org, "worker_id": body.worker_id,
         "column_key": body.column_key, "deleted_at": None},
        {"_id": 0},
    )

    if body.clear:
        if existing:
            await db.worker_certifications.update_one(
                {"id": existing["id"]},
                {"$set": {"deleted_at": _now_iso(), "updated_at": _now_iso(),
                          "updated_by": user["id"]}},
            )
        return {"ok": True, "cleared": True}

    # Validate ISO date.
    if body.expiry_date:
        try:
            date.fromisoformat(body.expiry_date)
        except Exception:
            raise HTTPException(400, "expiry_date must be ISO YYYY-MM-DD")

    patch: dict = {"updated_at": _now_iso(), "updated_by": user["id"]}
    if body.header is not None:        patch["name"] = body.header
    if body.category is not None:      patch["category"] = body.category
    if body.expiry_date is not None:   patch["expiry_date"] = body.expiry_date or None
    if body.not_held is not None:      patch["not_held"] = body.not_held
    if body.held_no_expiry is not None: patch["held_no_expiry"] = body.held_no_expiry
    if body.doc_file_id is not None:   patch["doc_file_id"] = body.doc_file_id
    # If the user supplies a date, automatically un-flag not_held.
    if patch.get("expiry_date"):       patch["not_held"] = False
    # Setting not_held=True wipes the date.
    if patch.get("not_held"):          patch["expiry_date"] = None
    patch["import_confidence"] = "manual"

    if existing:
        await db.worker_certifications.update_one(
            {"id": existing["id"]}, {"$set": patch},
        )
        cert_id = existing["id"]
    else:
        doc = {
            "id": _new_id(), "org_id": org, "worker_id": body.worker_id,
            "column_key": body.column_key,
            "name": body.header or body.column_key,
            "category": body.category or "competency",
            "issuer": "", "issue_date": None,
            "expiry_date": None, "not_held": False, "held_no_expiry": False,
            "doc_file_id": None, "doc_folder_id": None,
            "source": "manual_matrix_edit", "import_confidence": "manual",
            "notes": "", "created_by": user["id"],
            "created_at": _now_iso(), "deleted_at": None,
        }
        doc.update(patch)
        await db.worker_certifications.insert_one(doc)
        cert_id = doc["id"]

    out = await db.worker_certifications.find_one({"id": cert_id}, {"_id": 0})
    return {"ok": True, "cert_id": cert_id, "status": _status_for(out)}


# ────────────────── export.xlsx ──────────────────

from fastapi.responses import StreamingResponse  # noqa: E402


# ────────────────── print endpoint ──────────────────

class PrintIn(BaseModel):
    worker_ids: list[str]
    layout: str = "a4_landscape"        # a4_portrait | a4_landscape | a3_landscape
    include_cover: bool = True
    include_legend: bool = True
    include_raw: bool = False
    include_last_updated: bool = False
    combined: bool = True
    # download → forces browser save (Content-Disposition: attachment).
    # inline   → renders in-app iframe via PdfPreviewModal (Content-Disposition: inline).
    mode: str = "download"


_STATUS_LABEL = {
    "current": "Current", "expiring": "Expiring", "expired": "Expired",
    "not_held": "Not held", "held_no_expiry": "Held", "invalid_date": "Invalid",
    "unknown": "—",
}
_STATUS_RGB = {
    "current":        (0.85, 0.93, 0.87),
    "expiring":       (0.99, 0.95, 0.78),
    "expired":        (0.98, 0.89, 0.91),
    "not_held":       (0.94, 0.94, 0.94),
    "held_no_expiry": (0.90, 0.93, 0.97),
    "invalid_date":   (0.98, 0.89, 0.91),
    "unknown":        (0.97, 0.97, 0.97),
}


@router.post("/print")
async def print_inductions(body: PrintIn, user: dict = Depends(get_current_user)):
    """Generate a PDF for one or many workers' induction status."""
    if user.get("role") not in {"admin", "manager", "hseq_lead"}:
        raise HTTPException(403, "Only admin/manager/HSEQ can print inductions")
    if not body.worker_ids:
        raise HTTPException(400, "worker_ids must not be empty")

    from reportlab.lib.pagesizes import A4, A3, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table, TableStyle,
                                    Spacer, PageBreak)

    layout = body.layout
    page_size = (
        landscape(A3) if layout == "a3_landscape" else
        landscape(A4) if layout == "a4_landscape" else
        A4
    )

    # Pull the full matrix once (reuses the live computation including chip).
    data = await induction_matrix(user)
    by_id = {r["id"]: r for r in data["rows"]}
    cols = data["columns"]
    cols_by_cat = {"site_induction": [], "competency": [], "license": []}
    for c in cols:
        cols_by_cat.setdefault(c["category"], []).append(c)

    workers_to_print = [by_id[wid] for wid in body.worker_ids if wid in by_id]
    if not workers_to_print:
        raise HTTPException(404, "None of the supplied worker_ids matched")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=page_size,
                            leftMargin=18*mm, rightMargin=18*mm,
                            topMargin=14*mm, bottomMargin=14*mm,
                            title="Paneltec Civil — Inductions")
    styles = getSampleStyleSheet()
    h_org = ParagraphStyle("h_org", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=10, textColor=colors.HexColor("#2C6BFF"),
        leading=12, spaceAfter=2)
    h_name = ParagraphStyle("h_name", parent=styles["Heading1"],
        fontName="Helvetica-Bold", fontSize=20, leading=24, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9,
        textColor=colors.HexColor("#475569"), spaceAfter=8)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"],
        fontName="Helvetica-Bold", fontSize=12, textColor=colors.HexColor("#0f172a"),
        leading=14, spaceBefore=10, spaceAfter=4)

    elements = []

    # Cover page (combined mode + opted in).
    if body.combined and body.include_cover and len(workers_to_print) > 1:
        elements.append(Paragraph("PANELTEC CIVIL", h_org))
        elements.append(Paragraph("Inductions Print Pack", h_name))
        elements.append(Paragraph(
            f"{len(workers_to_print)} worker(s) · generated {datetime.now(timezone.utc).strftime('%d %b %Y · %H:%M UTC')}",
            sub))
        cover_rows = [["#", "Worker", "Company", "Overall"]]
        for i, w in enumerate(workers_to_print, 1):
            cover_rows.append([str(i), w["name"], w.get("company") or "—",
                               _STATUS_LABEL.get(w.get("chip"), w.get("chip") or "—")])
        t = Table(cover_rows, colWidths=[12*mm, 75*mm, 60*mm, 30*mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f1f5f9")),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#e2e8f0")),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    # Per-worker pages.
    for idx, w in enumerate(workers_to_print):
        elements.append(Paragraph("PANELTEC CIVIL · INDUCTIONS", h_org))
        elements.append(Paragraph(w["name"], h_name))
        elements.append(Paragraph(
            f"{w.get('company') or '—'} · status: {_STATUS_LABEL.get(w.get('chip'), '—')}"
            f" · printed {datetime.now(timezone.utc).strftime('%d %b %Y')}", sub))

        # Per-category sections.
        cats = [("site_induction", "Site Inductions"),
                ("competency", "Competencies"),
                ("license", "Licences")]
        any_cell = False
        for cat_key, cat_label in cats:
            cat_cols = cols_by_cat.get(cat_key, [])
            if not cat_cols:
                continue
            elements.append(Paragraph(cat_label, h2))
            header = ["Item", "Status", "Expiry"]
            if body.include_raw:
                header.append("Source")
            if body.include_last_updated:
                header.append("Updated")
            table_rows = [header]
            cell_colors = []
            for c in cat_cols:
                cell = w["cells"].get(c["column_key"])
                if cell:
                    any_cell = True
                    status = cell.get("status") or "unknown"
                    expiry = cell.get("expiry_date") or "—"
                    if cell.get("not_held"): expiry = "—"
                    src = cell.get("import_confidence") or "manual"
                    upd = "—"
                else:
                    status, expiry, src, upd = "unknown", "—", "—", "—"
                row = [c["header"], _STATUS_LABEL[status], expiry]
                if body.include_raw: row.append(src)
                if body.include_last_updated: row.append(upd)
                table_rows.append(row)
                cell_colors.append(_STATUS_RGB.get(status, (1, 1, 1)))
            col_widths = [70*mm, 28*mm, 28*mm]
            if body.include_raw: col_widths.append(28*mm)
            if body.include_last_updated: col_widths.append(28*mm)
            t = Table(table_rows, colWidths=col_widths, repeatRows=1)
            ts = TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f1f5f9")),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 9),
                ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#e2e8f0")),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                ("LEFTPADDING", (0,0), (-1,-1), 5),
                ("RIGHTPADDING", (0,0), (-1,-1), 5),
                ("TOPPADDING", (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ])
            for i, rgb in enumerate(cell_colors, start=1):
                ts.add("BACKGROUND", (1, i), (1, i), colors.Color(*rgb))
            t.setStyle(ts)
            elements.append(t)

        # Access section.
        a = w.get("access") or {}
        if any(a.get(k) for k in ("vehicle", "building_key", "gate_key")) or a.get("extras"):
            elements.append(Paragraph("Access", h2))
            ar = [["Item", "Held"]]
            ar.append(["Vehicle", "Yes" if a.get("vehicle") else "—"])
            ar.append(["Building key", "Yes" if a.get("building_key") else "—"])
            ar.append(["Gate key", "Yes" if a.get("gate_key") else "—"])
            if a.get("extras"):
                ar.append(["Notes", a["extras"]])
            t = Table(ar, colWidths=[70*mm, 56*mm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f1f5f9")),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 9),
                ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#e2e8f0")),
                ("LEFTPADDING", (0,0), (-1,-1), 5), ("RIGHTPADDING", (0,0), (-1,-1), 5),
                ("TOPPADDING", (0,0), (-1,-1), 4), ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ]))
            elements.append(t)

        if not any_cell:
            elements.append(Spacer(1, 4*mm))
            elements.append(Paragraph(
                "<i>No induction records on file for this worker.</i>",
                ParagraphStyle("empty", parent=styles["Normal"], fontSize=10,
                               textColor=colors.HexColor("#94a3b8"))))

        if body.include_legend:
            elements.append(Spacer(1, 6*mm))
            elements.append(Paragraph(
                "<font color='#94a3b8' size='8'>Legend: "
                "<b>Current</b> · <b>Expiring</b> within 30 days · <b>Expired</b> past expiry · "
                "<b>Not held</b> · <b>Held</b> (no expiry on file) · <b>Invalid</b> date format."
                "</font>", styles["Normal"]))

        # Combined mode → page break between workers (except last).
        if body.combined and idx < len(workers_to_print) - 1:
            elements.append(PageBreak())
        elif not body.combined and idx < len(workers_to_print) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    buf.seek(0)
    fname = "paneltec-inductions.pdf" if body.combined else f"paneltec-inductions-{workers_to_print[0]['name'].replace(' ','_')}.pdf"
    disposition = "inline" if body.mode == "inline" else "attachment"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f'{disposition}; filename="{fname}"'})


# ────────────────── export.xlsx ──────────────────


@router.get("/export.xlsx")
async def export_matrix(user: dict = Depends(get_current_user)):
    """Round-trip the live inductions matrix back to .xlsx. Sheet 1
    'Inductions' has Worker × Column with ISO dates / not held flags.
    Sheet 2 'Accessibilities' lists per-worker vehicle / key access."""
    if user.get("role") not in {"admin", "manager", "hseq_lead"}:
        raise HTTPException(403, "Only admin/manager/HSEQ can export inductions")

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    data = await induction_matrix(user)
    cols = data["columns"]
    rows = data["rows"]

    wb = openpyxl.Workbook()

    # Sheet 1 — Inductions.
    ws = wb.active
    ws.title = "Inductions"
    header = ["Worker", "Company", "Email"] + [c["header"] for c in cols]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        line = [r["name"], r.get("company", ""), r.get("email", "")]
        for c in cols:
            cell_state = r["cells"].get(c["column_key"])
            if not cell_state:
                line.append("")
            elif cell_state.get("not_held"):
                line.append("Not held")
            elif cell_state.get("held_no_expiry"):
                line.append("Held")
            else:
                line.append(cell_state.get("expiry_date") or "")
        ws.append(line)
    for col_idx, h in enumerate(header, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, min(40, len(str(h)) + 4))

    # Sheet 2 — Accessibilities.
    ws2 = wb.create_sheet("Accessibilities")
    ws2.append(["Worker", "Vehicle", "Building Key", "Gate Key", "Extras"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in rows:
        a = r["access"]
        ws2.append([
            r["name"],
            "Y" if a.get("vehicle") else "",
            "Y" if a.get("building_key") else "",
            "Y" if a.get("gate_key") else "",
            a.get("extras", ""),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="paneltec-inductions-matrix.xlsx"'},
    )


# ────────────────── Phase 3.12 — Induction Card popup endpoints ──────────────────
#
# These hang off `/workers/{worker_id}/inductions/...` (distinct from the existing
# `/workers/inductions/*` router) and serve a single, normalised induction record
# for the in-modal viewer. They are thin wrappers around the same
# `worker_certifications` collection used by the live matrix.

from fastapi import File, UploadFile, Response  # noqa: E402
from pathlib import Path  # noqa: E402
import uuid as _uuid  # noqa: E402

card_router = APIRouter(prefix="/workers", tags=["workers-inductions-card"])

_CARD_WRITE = {"admin", "manager", "hseq_lead"}


def _ind_status(cert: dict) -> str:
    """Same chip logic used by the matrix — kept here so the card payload
    is self-contained and doesn't require a full matrix fetch."""
    if cert.get("not_held"):
        return "not_held"
    exp = cert.get("expiry_date")
    if not exp:
        return "held_no_expiry" if cert.get("held_no_expiry") else "unknown"
    try:
        d = date.fromisoformat(exp)
    except Exception:
        return "invalid_date"
    today = date.today()
    if d < today:
        return "expired"
    delta = (d - today).days
    if delta <= 30:
        return "expiring"
    if delta <= 90:
        return "expiring_90"
    return "current"


def _serialise_induction(cert: dict) -> dict:
    return {
        "id": cert["id"],
        "worker_id": cert["worker_id"],
        "name": cert.get("name") or "",
        "type": cert.get("category") or "competency",
        "category": cert.get("category") or "competency",
        "column_key": cert.get("column_key"),
        "issuer": cert.get("issuer") or "",
        "issue_date": cert.get("issue_date"),
        "expiry_date": cert.get("expiry_date"),
        "not_held": bool(cert.get("not_held")),
        "held_no_expiry": bool(cert.get("held_no_expiry")),
        "status": cert.get("status_override") or _ind_status(cert),
        "status_override": cert.get("status_override"),
        "notes": cert.get("notes") or "",
        "doc_file_id": cert.get("doc_file_id"),
        "doc_folder_id": cert.get("doc_folder_id"),
        "raw_input": cert.get("raw_input"),
        "import_confidence": cert.get("import_confidence"),
        "updated_at": cert.get("updated_at"),
        "created_at": cert.get("created_at"),
    }


async def _worker_or_404(worker_id: str, org_id: str) -> dict:
    w = await db.workers.find_one(
        {"id": worker_id, "org_id": org_id, "deleted_at": None}, {"_id": 0},
    )
    if not w:
        raise HTTPException(404, "Worker not found")
    return w


def _can_read_own(user: dict, worker: dict) -> bool:
    """A worker user may read inductions for the worker record whose email
    matches their login email (case-insensitive). Simpro id fallback too."""
    if user.get("role") in _CARD_WRITE:
        return True
    if user.get("role") != "worker":
        return False
    u_email = (user.get("email") or "").strip().lower()
    w_email = (worker.get("email") or "").strip().lower()
    if u_email and u_email == w_email:
        return True
    if user.get("simpro_employee_id") and worker.get("simpro_employee_id"):
        return str(user["simpro_employee_id"]) == str(worker["simpro_employee_id"])
    return False


async def _induction_or_404(induction_id: str, worker_id: str, org_id: str) -> dict:
    """Returns the cert document; 404 if it doesn't belong to this worker
    or this org. Don't leak existence across worker boundaries."""
    cert = await db.worker_certifications.find_one(
        {"id": induction_id, "org_id": org_id, "deleted_at": None},
        {"_id": 0},
    )
    if not cert or cert.get("worker_id") != worker_id:
        raise HTTPException(404, "Induction not found")
    return cert


class IndCreateIn(BaseModel):
    name: str
    type: Optional[str] = None          # site_induction | competency | license
    issuer: Optional[str] = ""
    issue_date: Optional[str] = None
    expiry_date: Optional[str] = None
    notes: Optional[str] = ""
    not_held: Optional[bool] = False
    held_no_expiry: Optional[bool] = False
    column_key: Optional[str] = None


class IndPatch(BaseModel):
    issuer: Optional[str] = None
    issue_date: Optional[str] = None
    expiry_date: Optional[str] = None
    notes: Optional[str] = None
    not_held: Optional[bool] = None
    held_no_expiry: Optional[bool] = None
    name: Optional[str] = None
    status_override: Optional[str] = None  # admin-only — bypasses date logic
    # Sentinel: pass `clear_status_override: true` to remove an override.
    clear_status_override: Optional[bool] = False


@card_router.get("/{worker_id}/inductions/{induction_id}")
async def get_induction(worker_id: str, induction_id: str, user: dict = Depends(get_current_user)):
    worker = await _worker_or_404(worker_id, user["org_id"])
    if not _can_read_own(user, worker):
        # Worker users probing other workers' records get 404 (not 403) so we
        # don't leak existence of those records — same shape as a genuine
        # not-found. Non-worker authenticated roles that lack read access
        # (supervisor/auditor without permission) keep the 403 since they're
        # legitimately internal users and the response shape is informative.
        if user.get("role") == "worker":
            raise HTTPException(404, "Induction not found")
        raise HTTPException(403, "Permission denied")
    cert = await _induction_or_404(induction_id, worker_id, user["org_id"])
    return _serialise_induction(cert)


@card_router.patch("/{worker_id}/inductions/{induction_id}")
async def patch_induction(
    worker_id: str, induction_id: str, body: IndPatch,
    user: dict = Depends(get_current_user),
):
    if user.get("role") not in _CARD_WRITE:
        raise HTTPException(403, "Only admin/manager/HSEQ can edit inductions")
    await _worker_or_404(worker_id, user["org_id"])
    cert = await _induction_or_404(induction_id, worker_id, user["org_id"])

    patch = {"updated_at": _now_iso(), "updated_by": user["id"]}
    if body.name is not None:        patch["name"] = body.name.strip()
    if body.issuer is not None:      patch["issuer"] = body.issuer.strip()
    if body.notes is not None:       patch["notes"] = body.notes.strip()
    if body.issue_date is not None:
        if body.issue_date and not _is_iso(body.issue_date):
            raise HTTPException(400, "issue_date must be ISO YYYY-MM-DD")
        patch["issue_date"] = body.issue_date or None
    if body.expiry_date is not None:
        if body.expiry_date and not _is_iso(body.expiry_date):
            raise HTTPException(400, "expiry_date must be ISO YYYY-MM-DD")
        patch["expiry_date"] = body.expiry_date or None
    if body.not_held is not None:        patch["not_held"] = bool(body.not_held)
    if body.held_no_expiry is not None:  patch["held_no_expiry"] = bool(body.held_no_expiry)
    if body.clear_status_override:
        patch["status_override"] = None
    elif body.status_override is not None:
        if user.get("role") != "admin":
            raise HTTPException(403, "Only admin may override status")
        patch["status_override"] = body.status_override
    # If date is set, ensure not_held is false (consistency).
    if patch.get("expiry_date"):
        patch["not_held"] = False

    await db.worker_certifications.update_one({"id": cert["id"]}, {"$set": patch})
    out = await db.worker_certifications.find_one({"id": cert["id"]}, {"_id": 0})
    return _serialise_induction(out)


def _is_iso(value: str) -> bool:
    try:
        date.fromisoformat(value)
        return True
    except Exception:
        return False


@card_router.post("/{worker_id}/inductions", status_code=201)
async def create_induction(
    worker_id: str, body: IndCreateIn, user: dict = Depends(get_current_user),
):
    if user.get("role") not in _CARD_WRITE:
        raise HTTPException(403, "Only admin/manager/HSEQ can add inductions")
    await _worker_or_404(worker_id, user["org_id"])

    name = (body.name or "").strip()
    if not name:
        raise HTTPException(400, "name is required")
    column_key = body.column_key or _column_key(name)
    category = body.type or _category_for_header(name)
    for d in (body.issue_date, body.expiry_date):
        if d and not _is_iso(d):
            raise HTTPException(400, f"date must be ISO YYYY-MM-DD: {d}")

    # If a cert with this column_key already exists for the worker, reject —
    # the matrix is keyed on (worker_id, column_key) so duplicates would
    # silently overwrite. Frontend should call PATCH instead.
    dup = await db.worker_certifications.find_one(
        {"org_id": user["org_id"], "worker_id": worker_id,
         "column_key": column_key, "deleted_at": None},
        {"_id": 0, "id": 1},
    )
    if dup:
        raise HTTPException(409, f"Induction '{name}' already exists for this worker")

    doc = {
        "id": _new_id(), "org_id": user["org_id"], "worker_id": worker_id,
        "name": name,
        "category": category,
        "column_key": column_key,
        "issuer": (body.issuer or "").strip(),
        "issue_date": body.issue_date,
        "expiry_date": body.expiry_date,
        "not_held": bool(body.not_held),
        "held_no_expiry": bool(body.held_no_expiry),
        "doc_file_id": None,
        "doc_folder_id": None,
        "notes": (body.notes or "").strip(),
        "source": "manual_card_add",
        "import_confidence": "manual",
        "created_by": user["id"],
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
        "deleted_at": None,
    }
    await db.worker_certifications.insert_one(doc)
    return _serialise_induction(doc)


@card_router.post("/{worker_id}/inductions/{induction_id}/file", status_code=201)
async def upload_induction_file(
    worker_id: str, induction_id: str,
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """Replace (or set) the certificate document attached to an induction.
    Reuses Document Library storage + smart folder routing (same code path
    used by `POST /workers/{wid}/certifications/upload`)."""
    if user.get("role") not in _CARD_WRITE:
        raise HTTPException(403, "Only admin/manager/HSEQ can upload induction docs")
    worker = await _worker_or_404(worker_id, user["org_id"])
    cert = await _induction_or_404(induction_id, worker_id, user["org_id"])

    # Lazy import to avoid heavy startup cost if Document Library not used yet.
    from document_library import (
        MAX_FILE_BYTES, UPLOAD_DIR, _safe_ext, _stub_ai_tags,
    )
    from worker_certifications import (
        _match_folder_name, _resolve_seed_folder, _find_or_create_worker_subfolder,
    )

    ext = _safe_ext(file.filename)
    if not ext:
        raise HTTPException(400, "Unsupported file type")

    cert_name = cert.get("name") or Path(file.filename or "").stem or "Induction"
    seed_name = _match_folder_name(cert_name)
    seed_folder = await _resolve_seed_folder(user["org_id"], seed_name, user["id"])
    sub_folder = await _find_or_create_worker_subfolder(seed_folder, worker, user["id"])

    folder_dir = UPLOAD_DIR / sub_folder["id"]
    folder_dir.mkdir(parents=True, exist_ok=True)
    stored_name = f"{_uuid.uuid4().hex}{ext}"
    target = folder_dir / stored_name

    size = 0
    with target.open("wb") as out:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk: break
            size += len(chunk)
            if size > MAX_FILE_BYTES:
                out.close(); target.unlink(missing_ok=True)
                raise HTTPException(400, "Exceeds 50 MB limit")
            out.write(chunk)

    worker_label = f"{worker.get('first_name', '')} {worker.get('last_name', '')}".strip() or "(unnamed)"
    file_doc = {
        "id": _new_id(),
        "org_id": user["org_id"],
        "folder_id": sub_folder["id"],
        "filename": file.filename or stored_name,
        "stored_name": stored_name,
        "mime": file.content_type or "application/octet-stream",
        "size": size,
        "file_url": f"/api/files/document_library/{sub_folder['id']}/{stored_name}",
        "uploaded_by": user["id"],
        "uploaded_by_name": user.get("name") or user.get("email"),
        "uploaded_at": _now_iso(),
        "updated_at": _now_iso(),
        "ai_tags": _stub_ai_tags(file.filename or stored_name) + [f"worker:{worker_label}", f"induction:{cert_name}"],
        "uploaded_via": "induction_card",
        "worker_id": worker_id,
        "worker_name": worker_label,
        "seed_folder": seed_folder["name"],
        "deleted_at": None,
    }
    await db.doc_files.insert_one(file_doc)

    await db.worker_certifications.update_one(
        {"id": cert["id"]},
        {"$set": {"doc_file_id": file_doc["id"], "doc_folder_id": sub_folder["id"],
                  "updated_at": _now_iso(), "updated_by": user["id"]}},
    )
    updated = await db.worker_certifications.find_one({"id": cert["id"]}, {"_id": 0})
    return {"ok": True, "induction": _serialise_induction(updated),
            "file": {"id": file_doc["id"], "filename": file_doc["filename"]}}


@card_router.delete("/{worker_id}/inductions/{induction_id}", status_code=204)
async def delete_induction(
    worker_id: str, induction_id: str, user: dict = Depends(get_current_user),
):
    if user.get("role") != "admin":
        raise HTTPException(403, "Only admin can delete inductions")
    await _worker_or_404(worker_id, user["org_id"])
    cert = await _induction_or_404(induction_id, worker_id, user["org_id"])
    await db.worker_certifications.update_one(
        {"id": cert["id"]},
        {"$set": {"deleted_at": _now_iso(), "updated_at": _now_iso(),
                  "updated_by": user["id"]}},
    )
    return Response(status_code=204)
